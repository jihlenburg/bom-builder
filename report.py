"""Writers for BOM output formats and shared presentation helpers.

The report layer is responsible only for formatting already-computed pricing
data. It does not recalculate totals or resolve parts. Each writer consumes the
same :class:`models.BomSummary` object so CSV, Excel, JSON, and console output
all agree on totals and error counts.
"""

import csv
import json
from dataclasses import dataclass
from pathlib import Path
import re
from typing import Any, Callable

from console import console, Panel
from manufacturer_packaging import packaging_kind_from_text
from models import BomSummary, PricedPart

@dataclass(frozen=True)
class ColumnSpec:
    """Definition of one output column in tabular report formats.

    Attributes
    ----------
    header:
        Human-readable column title.
    accessor:
        Callable that extracts or formats the corresponding value from a
        :class:`PricedPart`.
    """

    header: str
    accessor: Callable[[PricedPart], Any]


def _packaging_text(part: PricedPart) -> str:
    """Return lower-cased packaging context text for batch-plan heuristics."""
    return " ".join(
        text
        for text in [part.pricing_strategy, part.packaging_mode, part.package_type]
        if text
    ).lower()


def _single_purchase_leg(part: PricedPart):
    """Return the only recorded purchase leg, or ``None`` for mixed plans."""
    return part.purchase_legs[0] if len(part.purchase_legs) == 1 else None


def _batch_kind_from_text(packaging_text: str) -> str | None:
    """Infer the packaging noun to use in batch-plan formatting.

    Delegates to :func:`manufacturer_packaging.packaging_kind_from_text` for
    the actual keyword matching logic.
    """
    return packaging_kind_from_text(packaging_text)


def _order_batch_details(part: PricedPart) -> tuple[int | None, str | None]:
    """Return the selected order-batch size and label, when it is inferable."""
    single_leg = _single_purchase_leg(part)
    if (
        single_leg is not None
        and single_leg.order_batch_quantity is not None
        and single_leg.order_batch_count is not None
    ):
        packaging_text = " ".join(
            text for text in [single_leg.packaging_mode, single_leg.package_type] if text
        )
        return single_leg.order_batch_quantity, _batch_kind_from_text(packaging_text)

    purchased_quantity = part.purchased_quantity or 0
    if purchased_quantity <= 0:
        return None, None

    packaging_text = _packaging_text(part)
    full_reel_quantity = part.full_reel_quantity or 0
    if full_reel_quantity > 1 and purchased_quantity % full_reel_quantity == 0:
        reel_selected = (
            "full reel" in packaging_text
            or (
                "reel" in packaging_text
                and "cut tape" not in packaging_text
                and "mousereel" not in packaging_text
            )
            or part.order_multiple in (None, 0, 1, full_reel_quantity)
        )
        if reel_selected:
            return full_reel_quantity, "reel"

    order_multiple = part.order_multiple or 0
    if order_multiple > 1 and purchased_quantity % order_multiple == 0:
        return order_multiple, "batch"

    minimum_order_quantity = part.minimum_order_quantity or 0
    if minimum_order_quantity > 1 and purchased_quantity % minimum_order_quantity == 0:
        return minimum_order_quantity, "lot"

    return None, None


def _order_batch_quantity(part: PricedPart) -> int | str:
    """Return the inferred batch/reel size used for the selected buy."""
    batch_quantity, _ = _order_batch_details(part)
    return batch_quantity if batch_quantity is not None else ""


def _order_batch_count(part: PricedPart) -> int | str:
    """Return the inferred number of batches/reels used for the selected buy."""
    single_leg = _single_purchase_leg(part)
    if single_leg is not None and single_leg.order_batch_count is not None:
        return single_leg.order_batch_count

    batch_quantity, _ = _order_batch_details(part)
    purchased_quantity = part.purchased_quantity or 0
    if batch_quantity is None or purchased_quantity <= 0:
        return ""
    return purchased_quantity // batch_quantity


def _order_plan(part: PricedPart) -> str:
    """Return a compact human-readable order plan for CSV/Excel output."""
    if part.order_plan:
        return part.order_plan

    batch_quantity, batch_kind = _order_batch_details(part)
    purchased_quantity = part.purchased_quantity or 0
    if batch_quantity is None or batch_kind is None or purchased_quantity <= 0:
        return ""

    batch_count = purchased_quantity // batch_quantity
    batch_noun = {
        "reel": "reel" if batch_count == 1 else "reels",
        "batch": "batch" if batch_count == 1 else "batches",
        "lot": "lot" if batch_count == 1 else "lots",
    }[batch_kind]
    return f"{batch_count} {batch_noun} x {batch_quantity}"


def _available_quantity(part: PricedPart) -> int | str:
    """Return the parsed numeric stock quantity when availability is explicit."""
    if not part.availability:
        return ""

    match = re.search(
        r"([\d,.]+)\s*(?:in stock|available)",
        part.availability,
        flags=re.IGNORECASE,
    )
    if match is None:
        return ""

    token = match.group(1).replace(",", "").replace(".", "")
    try:
        return int(token)
    except ValueError:
        return ""


def _shortage_quantity(part: PricedPart) -> int | str:
    """Return the uncovered portion of the selected order quantity, if known."""
    available_quantity = _available_quantity(part)
    if available_quantity == "" or part.purchased_quantity is None:
        return ""
    return max(part.purchased_quantity - available_quantity, 0)


def _line_status(part: PricedPart) -> str:
    """Return the operational status shown at the front of the sheet."""
    if part.lookup_error:
        return "ERROR"
    if part.extended_price is None:
        return "NO PRICE"
    if part.review_required:
        return "REVIEW"
    shortage_quantity = _shortage_quantity(part)
    if shortage_quantity not in ("", 0):
        return "SHORT"
    return "OK"


# Column definitions — order here determines output column order.
COLUMNS: tuple[ColumnSpec, ...] = (
    ColumnSpec("Status", _line_status),
    ColumnSpec("Distributor", lambda p: p.distributor or ""),
    ColumnSpec("Distributor PN", lambda p: p.distributor_part_number or ""),
    ColumnSpec("Manufacturer", lambda p: p.manufacturer),
    ColumnSpec("Manufacturer PN", lambda p: p.manufacturer_part_number or ""),
    ColumnSpec("Part Number", lambda p: p.part_number),
    ColumnSpec("Description", lambda p: p.description or ""),
    ColumnSpec("Package", lambda p: p.package or ""),
    ColumnSpec("Pins", lambda p: p.pins if p.pins is not None else ""),
    ColumnSpec("Reference", lambda p: p.reference or ""),
    ColumnSpec("Qty/Unit", lambda p: p.quantity_per_unit),
    ColumnSpec("Build Need", lambda p: p.total_quantity),
    ColumnSpec("Order Qty", lambda p: p.purchased_quantity if p.purchased_quantity is not None else ""),
    ColumnSpec("Shortage Qty", _shortage_quantity),
    ColumnSpec("Overbuy Qty", lambda p: p.surplus_quantity if p.surplus_quantity is not None else ""),
    ColumnSpec("Order Plan", _order_plan),
    ColumnSpec("Order Batch Qty", _order_batch_quantity),
    ColumnSpec("Order Batch Count", _order_batch_count),
    ColumnSpec("Pricing Strategy", lambda p: p.pricing_strategy or ""),
    ColumnSpec("Price Break Qty", lambda p: p.price_break_quantity or ""),
    ColumnSpec("Package Type", lambda p: p.package_type or ""),
    ColumnSpec("Packaging Mode", lambda p: p.packaging_mode or ""),
    ColumnSpec("MOQ", lambda p: p.minimum_order_quantity if p.minimum_order_quantity is not None else ""),
    ColumnSpec("Order Multiple", lambda p: p.order_multiple if p.order_multiple is not None else ""),
    ColumnSpec("Full Reel Qty", lambda p: p.full_reel_quantity if p.full_reel_quantity is not None else ""),
    ColumnSpec("Available Now", _available_quantity),
    ColumnSpec("Availability Detail", lambda p: p.availability or ""),
    ColumnSpec("Unit Price", lambda p: f"{p.unit_price:.4f}" if p.unit_price is not None else ""),
    ColumnSpec("Extended Price", lambda p: f"{p.extended_price:.2f}" if p.extended_price is not None else ""),
    ColumnSpec("Currency", lambda p: p.currency or ""),
    ColumnSpec("Match Method", lambda p: p.match_method.value if p.match_method else ""),
    ColumnSpec("Resolution Source", lambda p: p.resolution_source or ""),
    ColumnSpec("Packaging Source", lambda p: p.packaging_source or ""),
    ColumnSpec("Candidates", lambda p: p.match_candidates if p.match_candidates is not None else ""),
    ColumnSpec("Errors", lambda p: p.lookup_error or ""),
)

HEADER_ROW = [column.header for column in COLUMNS]
EXTENDED_PRICE_INDEX = next(
    i for i, column in enumerate(COLUMNS) if column.header == "Extended Price"
)


def _prepare_output_path(output: Path) -> None:
    """Ensure the destination directory exists before writing a report."""
    output.parent.mkdir(parents=True, exist_ok=True)


def _part_to_row(p: PricedPart) -> list:
    """Convert a priced part into one tabular output row."""
    return [column.accessor(p) for column in COLUMNS]


def _summary_row(label: str, value: str) -> list:
    """Create one footer row aligned with the extended-price column."""
    row = [label] + [""] * (len(COLUMNS) - 1)
    row[EXTENDED_PRICE_INDEX] = value
    return row


def _summary_rows(summary: BomSummary) -> list[list[str]]:
    """Return the footer rows shared by CSV and Excel outputs."""
    rows = [
        [],
        _summary_row("BOM COST / UNIT", f"{summary.cost_per_unit:.2f}"),
        [f"UNIQUE PARTS: {summary.total_parts}"],
    ]
    if summary.error_count:
        rows.append([f"LOOKUP ERRORS: {summary.error_count}"])
    return rows


def _print_write_status(output: Path, summary: BomSummary) -> None:
    """Print the common post-write status message shown by all writers."""
    cur = summary.currency or ""
    lines = [
        f"BOM written to [heading]{output}[/heading]",
        f"  {summary.total_parts} unique parts, BOM cost / unit: [price]{summary.cost_per_unit:.2f} {cur}[/price]",
    ]
    if summary.error_count:
        lines.append(f"  [review]{summary.error_count} part(s) had lookup errors[/review]")
    console.print("\n".join(lines))


def write_csv(
    parts: list[PricedPart], output: Path, summary: BomSummary
) -> None:
    """Write the priced BOM to CSV.

    Parameters
    ----------
    parts:
        Final priced part records to write.
    output:
        Destination file path.
    summary:
        Precomputed BOM summary shared across all writers.
    """
    _prepare_output_path(output)
    with output.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(HEADER_ROW)

        for p in parts:
            writer.writerow(_part_to_row(p))

        for row in _summary_rows(summary):
            writer.writerow(row)

    _print_write_status(output, summary)


def write_excel(
    parts: list[PricedPart], output: Path, summary: BomSummary
) -> None:
    """Write the priced BOM to an Excel workbook.

    When :mod:`openpyxl` is unavailable, the function falls back to CSV output
    rather than failing the entire run.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.worksheet.table import Table, TableStyleInfo
    except ImportError:
        console.print("[review]openpyxl not installed. Install with: pip install openpyxl[/review]")
        console.print("Falling back to CSV output.")
        write_csv(parts, output.with_suffix(".csv"), summary)
        return

    _prepare_output_path(output)
    wb = Workbook()
    ws = wb.active
    ws.title = "eBOM"

    ws.append(HEADER_ROW)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for p in parts:
        ws.append(_part_to_row(p))

    data_end_row = ws.max_row
    ws.freeze_panes = "G2"
    if data_end_row > 1:
        last_col_letter = ws.cell(row=1, column=len(HEADER_ROW)).column_letter
        table_ref = f"A1:{last_col_letter}{data_end_row}"
        table = Table(displayName="EBOMTable", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)
        ws.auto_filter.ref = table_ref

        status_column = HEADER_ROW.index("Status") + 1
        status_fills = {
            "OK": PatternFill(fill_type="solid", fgColor="E2F0D9"),
            "REVIEW": PatternFill(fill_type="solid", fgColor="FFF2CC"),
            "SHORT": PatternFill(fill_type="solid", fgColor="FCE4D6"),
            "NO PRICE": PatternFill(fill_type="solid", fgColor="F4CCCC"),
            "ERROR": PatternFill(fill_type="solid", fgColor="EA9999"),
        }
        for row in range(2, data_end_row + 1):
            status_cell = ws.cell(row=row, column=status_column)
            fill = status_fills.get(str(status_cell.value or ""))
            if fill is not None:
                status_cell.fill = fill

    for row in _summary_rows(summary):
        ws.append(row)
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    wb.save(output)
    _print_write_status(output, summary)


def write_json(
    parts: list[PricedPart], output: Path, summary: BomSummary
) -> None:
    """Write the priced BOM and summary metadata to JSON."""
    bom = {
        **summary.model_dump(),
        "parts": [p.model_dump(exclude_none=True) for p in parts],
    }

    _prepare_output_path(output)
    with output.open("w", encoding="utf-8") as f:
        json.dump(bom, f, indent=2, default=str)

    _print_write_status(output, summary)
