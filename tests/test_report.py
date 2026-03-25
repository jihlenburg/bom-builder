"""Tests for report writers."""

import csv
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

from models import BomSummary, PricedPart
from report import EXTENDED_PRICE_INDEX, HEADER_ROW, write_csv, write_excel, write_json


def _make_parts() -> list[PricedPart]:
    return [
        PricedPart(
            part_number="R1",
            manufacturer="Yageo",
            distributor="Mouser",
            distributor_part_number="603-R1",
            quantity_per_unit=2,
            total_quantity=200,
            required_quantity=200,
            purchased_quantity=12000,
            surplus_quantity=11800,
            order_multiple=3000,
            full_reel_quantity=3000,
            packaging_mode="Full Reel",
            pricing_strategy="full reel",
            availability="15000 In Stock",
            unit_price=0.05,
            extended_price=10.0,
            currency="EUR",
        )
    ]


class TestWriteCsv:
    def test_uses_generic_distributor_columns(self):
        assert "Status" in HEADER_ROW
        assert "Distributor" in HEADER_ROW
        assert "Distributor PN" in HEADER_ROW
        assert "Manufacturer PN" in HEADER_ROW
        assert "Build Need" in HEADER_ROW
        assert "Order Qty" in HEADER_ROW
        assert "Shortage Qty" in HEADER_ROW
        assert "Overbuy Qty" in HEADER_ROW
        assert "Order Batch Qty" in HEADER_ROW
        assert "Order Batch Count" in HEADER_ROW
        assert "Order Plan" in HEADER_ROW
        assert "Available Now" in HEADER_ROW
        assert "Availability Detail" in HEADER_ROW
        assert "Pricing Strategy" in HEADER_ROW
        assert "Packaging Mode" in HEADER_ROW
        assert "Full Reel Qty" in HEADER_ROW
        assert "Mouser PN" not in HEADER_ROW

    def test_creates_parent_directory_and_writes_summary(self, tmp_path):
        parts = _make_parts()
        summary = BomSummary.from_parts(parts, units=100)
        output = tmp_path / "nested" / "reports" / "bom.csv"

        write_csv(parts, output, summary)

        assert output.exists()
        with output.open(newline="", encoding="utf-8") as f:
            rows = list(csv.reader(f))

        cost_per_unit_row = next(row for row in rows if row and row[0] == "BOM COST / UNIT")
        header_index = {name: index for index, name in enumerate(rows[0])}
        data_row = rows[1]

        assert cost_per_unit_row[EXTENDED_PRICE_INDEX] == "0.10"
        assert data_row[header_index["Status"]] == "OK"
        assert data_row[header_index["Order Batch Qty"]] == "3000"
        assert data_row[header_index["Order Batch Count"]] == "4"
        assert data_row[header_index["Order Plan"]] == "4 reels x 3000"
        assert data_row[header_index["Available Now"]] == "15000"

    def test_excel_writer_adds_single_sheet_buyer_formatting(self, tmp_path):
        parts = _make_parts()
        summary = BomSummary.from_parts(parts, units=100)
        output = tmp_path / "bom.xlsx"

        write_excel(parts, output, summary)

        from openpyxl import load_workbook

        wb = load_workbook(output)
        ws = wb[wb.sheetnames[0]]

        assert ws.freeze_panes == "G2"
        assert ws.auto_filter.ref is not None
        assert len(ws.tables) == 1


class TestWriteJson:
    def test_creates_parent_directory_and_serializes_summary(self, tmp_path):
        parts = _make_parts()
        summary = BomSummary.from_parts(parts, units=100)
        output = tmp_path / "nested" / "reports" / "bom.json"

        write_json(parts, output, summary)

        assert output.exists()
        payload = json.loads(output.read_text(encoding="utf-8"))
        assert payload["total_cost"] == 10.0
        assert payload["parts"][0]["part_number"] == "R1"
